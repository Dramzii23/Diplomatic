import fitz  
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from PyQt5.QtGui import QPixmap, QFont, QPainter, QPdfWriter, QColor, QPen, QImage, QTransform
from PyQt5.QtWidgets import QApplication, QDialog, QSplashScreen, QFileDialog, QPushButton, QMessageBox, QFontComboBox, QDial, QLabel, QTableView
from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt, QTimer, QUrl, QSizeF, QMarginsF
from PyQt5.QtGui import QFontMetrics, QDesktopServices, QStandardItemModel, QStandardItem
from PyQt5 import QtWidgets, uic
from PyQt5.QtGui import QPainter, QImage
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtWidgets import QGraphicsDropShadowEffect
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QGraphicsScene
from PyQt5.QtGui import QPixmap
from pdf2image import convert_from_path
from openpyxl import load_workbook
import pandas as pd
import os, sys, datetime, shutil

# PyMuPDF

# from PyQt5.QtPrintSupport import QPdfWriter




# import pyglet
# import os

# https://stackoverflow.com/questions/31836104/pyinstaller-and-onefile-how-to-include-an-image-in-the-exe-file
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class MainWindow(QDialog):
    def __init__(self):
        # Call the constructor of the parent class (QMainWindow)
        super(MainWindow, self).__init__()

        # Load the user interface from the "mainwindow.ui" file
        loadUi(resource_path("ui_files\\mainwindow2.ui"), self)
        
                
        # Set image for imageLabel
        self.image_file = resource_path('img\\defaultImage-white.png')
        self.imageLabel.setPixmap(QPixmap(self.image_file))
        
        # Set image for imageLabelFBStory
        self.image_file_fbstory = resource_path('img\\BGFBStory.png')
        self.imageLabelFBStory.setPixmap(QPixmap(self.image_file_fbstory))

        # Set image for imageLabelFBStoryAward
        self.image_file_award = resource_path('img\\defaultImage-white.png')
        pixmap_award = QPixmap(self.image_file_award)
        transform = QTransform().rotate(-15)  # Rotate 90 degrees
        
        self.rotated_pixmap_award = pixmap_award.transformed(transform)
        self.imageLabelFBStoryAward.setPixmap(QPixmap(self.rotated_pixmap_award))
        
        # Create a shadow effect
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setXOffset(10)
        shadow.setYOffset(10)
        shadow.setColor(QColor("grey"))
        
        self.imageLabelFBStoryAward.setStyleSheet("""
            border: none;
            box-shadow: 10px 10px 5px grey;
            
        """)
        
        # Set image for imageLabelFBStoryProfile
        self.image_file_profile = resource_path('img\\profile_imgs\\Pablo Daniel Gutierrez (5).jpg')
        self.imageLabelFBStoryProfile.setPixmap(QPixmap(self.image_file_profile))
        transform = QTransform().rotate(8.5)
        self.imageLabelFBStoryProfile.setPixmap(QPixmap(self.image_file_profile).transformed(transform))
        # Add shadow and remove border for imageLabelFBStoryProfile
        self.imageLabelFBStoryProfile.setStyleSheet("""
            border: none;
            
           
        """)
        
        # Set image for imageLabelFBStoryWhite
        self.image_file_bg = resource_path('img\\profileImage-shadow-grey.png')
        self.imageLabelFBStoryWhite.setPixmap(QPixmap(self.image_file_bg))
        transform = QTransform().rotate(8.5)
        self.imageLabelFBStoryWhite.setPixmap(QPixmap(self.image_file_bg).transformed(transform))
        
        self.imageLabelFBStoryWhite.setStyleSheet("""
            border: none;
                     
        """)
 
        # Connect the clicked signal of the uploadImage QPushButton to the upload_image method
        # This means that when the uploadImage button is clicked, the upload_image method will be called
        self.uploadImage.clicked.connect(self.upload_image)

        # Connect the clicked signal of the exportPDF QPushButton to the export_pdf method
        # This means that when the exportPDF button is clicked, the export_pdf method will be called
        self.exportPDF.clicked.connect(self.export_pdf)
        
        self.exportMultiPDF.clicked.connect(self.export_multi_pdf)
        
        # Assuming subirExcelButton and tableViewExcel are members of self
        self.subirExcelButton.clicked.connect(self.handle_subirExcelButton_clicked)
        
        self.tableViewExcel.clicked.connect(self.handle_tableview_clicked)
        
        self.exportSubirFotoBoton.clicked.connect(self.upload_profile_image)
    # END OF __init__ METHOD        
# TEST OPENPYXL     

    def upload_profile_image(self):
        # Open a QFileDialog to select an image file
        # The getOpenFileName method returns a tuple where the first element is the selected file path
        # The second element is the selected filter, which we don't need, hence the underscore
        bgImage, _ = QFileDialog.getOpenFileName(self, "Open Image", self.image_file_profile, "Image Files (*.png *.jpg *.bmp)")
        if bgImage:
            self.image_file_profile = bgImage

        # If a file was selected (i.e., if self.image_file is not an empty string)
        if self.image_file_profile:
            # Create a QPixmap from the image file
            pixmap = QPixmap(self.image_file_profile)
            # Scale the QPixmap to fit the size of the imageLabel, while keeping the aspect ratio
            pixmap = pixmap.scaled(self.imageLabelFBStoryProfile.size(), Qt.KeepAspectRatio)
            # Set the QPixmap as the pixmap of the imageLabel
            self.imageLabelFBStoryProfile.setPixmap(pixmap)
            transform = QTransform().rotate(8.5)
            self.imageLabelFBStoryProfile.setPixmap(QPixmap(self.image_file_profile).transformed(transform))
            # Add shadow and remove border for imageLabelFBStoryProfile
            self.imageLabelFBStoryProfile.setStyleSheet("""
                border: none;
                
            """)
             # Define the destination folder and new filename
            destination_folder = "img/profile_imgs/"
            new_filename = os.path.basename(self.image_file_profile)  # get the original filename

        # Create the full destination path
            destination_path = os.path.join(destination_folder, new_filename)

            # If the source and destination filenames are the same, add a suffix to the filename
            suffix = 0
            while os.path.exists(destination_path) and os.path.basename(self.image_file_profile) == new_filename:
                suffix += 1
                new_filename = f"{os.path.splitext(os.path.basename(self.image_file_profile))[0]}_{suffix}{os.path.splitext(self.image_file_profile)[1]}"  # append suffix to original filename
                destination_path = os.path.join(destination_folder, new_filename)

            # Copy the uploaded file to the destination path
            shutil.copy(self.image_file_profile, destination_path)
    def handle_subirExcelButton_clicked_OPNPYXL(self):
        file = self._open_excel_file()
        if file:
            df, images = self._read_excel_file_with_images(file)
            self._populate_table_view(df)
            self._create_tmp_images(images)

    def _open_excel_file(self):
        file, _ = QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx)')
        return file

    def _read_excel_file_with_images(self, file):
        wb = load_workbook(filename=file)
        ws = wb.active

        data = []
        images = []

        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if cell.hyperlink:
                    image = self._read_image(cell.hyperlink.target)
                    images.append(image)
                    row_data.append(image)
                else:
                    row_data.append(cell.value)
            data.append(row_data)

        df = pd.DataFrame(data)
        return df, images

    def _read_image(self, image_path):
        # Implement this function to read an image from a file and return it in a format that you can use
        pass

    def _populate_table_view(self, df):
        model = QStandardItemModel(df.shape[0], df.shape[1])
        for row in df.iterrows():
            for col, value in enumerate(row[1]):
                item = QStandardItem(str(value))
                model.setItem(row[0], col, item)
        self.tableViewExcel.setModel(model)

    def _create_tmp_images(self, images):
        # imagesTMP = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        imagesTMP = "imagesTMP"
        os.makedirs(f"TMP/{imagesTMP}", exist_ok=True)
        for i, line in enumerate(images):
            self.labelName.setText(str(line))
            # self.createTMPimages(i, line, imagesTMP)
            try:
                self.createTMPimages(i, line, imagesTMP)
            except Exception as e:
                print(f"Error: {e}")
# END TESTS OPENPYXL
    def handle_subirExcelButton_clicked(self):
            # QMessageBox.information(self, "Title", "Hello 1") 
            file, _ = QFileDialog.getOpenFileName(self, 'Open Excel File', 'excel/', 'Excel Files (*.xlsx)')
            # QMessageBox.information(self, "Title", "Hello 2") 
            if file:
                df = pd.read_excel(file)
                # QMessageBox.information(self, "Title", "Hello 3") 
                model = QStandardItemModel(df.shape[0], df.shape[1])
                
                for row in df.iterrows():
                    for col, value in enumerate(row[1]):
                        
                        item = QStandardItem(str(value))
                        model.setItem(row[0], col, item)
                self.tableViewExcel.setModel(model)
                # Get the first column of the DataFrame as a list
                first_column = df.iloc[:, 0].tolist()

                # Join the list into a single string with newline characters between each item
                first_column_str = '\n'.join(str(item) for item in first_column)
                
               
                for i, line in enumerate(first_column_str.split('\n')):
                    
                    self.labelName.setText(line)
                    # self.export_pdf(str(i) + " " + line)
                
               # Create a new directory with the current date and time
                # imagesTMP = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                imagesTMP = "imagesTMP"
                
                os.makedirs(f"TMP/{imagesTMP}", exist_ok=True)

                for i, line in enumerate(first_column_str.split('\n')):
                    self.labelName.setText(line)
                    self.createTMPimages(i, line, imagesTMP)  # Pass the row number and content to createTMPimages
                    
    


    def convert_pdf_to_image(self, pdf_path, output_path):
        # Convert the PDF to a list of images
        images = convert_from_path(pdf_path, poppler_path=r"Poppler-V-24-02-00/Library/bin")
        # self.pdf2image.convert_from_path('path/to/pdf',poppler_path=r"path\to\poppler\bin")

        # Save the first page of the PDF as a PNG
        if images:
            images[0].save(output_path, 'PNG')
    
    def createTMPimages(self, row_number, row_content, imagesTMP):
        print("Creating TMP images")
        pixmap = self.imageLabel.grab()

        # Create a QPainter object
        painter = QPainter(pixmap)

        # Set the color and font of the text
        painter.setPen(QColor(255, 255, 255))  # White color
        painter.setFont(QFont('Arial', 20))  # Arial font, size 20

        # Draw the label name onto the pixmap at position (50, 50)
        painter.drawText(50, 50, self.labelName.text())

        # End the QPainter object
        painter.end()





        # Convert the pixmap to an image
        image = pixmap.toImage()

        # Replace any characters in row_content that are not valid in filenames
        valid_filename = "".join(c for c in row_content if c.isalnum() or c in (' ', '.', '_')).rstrip()

        # Save the image to a file in the new directory, named after the row content
        image.save(f"TMP/{imagesTMP}/{valid_filename}_{row_number}.png")
        
# ///////////////  Convert the pixmap to an image
        folder_path=f"TMP/{imagesTMP}"
        filename=f"{valid_filename}_{row_number}.pdf"

        self.export_pdf(None, folder_path, filename, preview=False)

        
        PDFtoExport = folder_path + "/" + filename
        PNGfromPDF = folder_path + "/" + f"{valid_filename}_{row_number}.png"
        
        # Corrected function call
        self.convert_pdf_to_image(str(PDFtoExport), str(PNGfromPDF))
        print("TMP images created")
        # Delete the PDF file after the PNG is created
        os.remove(PDFtoExport)
        
        
#///////////// Convert the pixmap to an image
        
        
    def handle_tableview_clicked(self, index):
        # Check if the selected cell is in the first column
        if index.column() == 0:
            # Get the model from the QTableView
            model = self.tableViewExcel.model()

            # Get the text from the selected cell
            cell_text = model.data(index, Qt.DisplayRole)
            
             # Get the row number
            row_number = index.row()           

            # Set the text of the QLabel
            self.labelName.setText(str(cell_text))
             # Set the text of the QLabel2
            self.labelName2.setText(str(cell_text))
            
            self.imageLabelFBStoryProfile2.setPixmap(QPixmap(self.image_file_profile))
        
        
            print(f"cell_text: {cell_text}")
            print(f"index: {row_number}")
            
            
            # Set image for imageLabelFBStoryAward
            self.image_file_award = resource_path(f'TMP\\imagesTMP\\{cell_text}_{row_number}.png')
            pixmap_award = QPixmap(self.image_file_award)
            transform = QTransform().rotate(-15)  # First rotate the pixmap
            print(f"self.image_file_award: {self.image_file_award}")
            
            # You need to first rotate the pixmap and then set it to the label
            self.rotated_pixmap_award = pixmap_award.transformed(transform)
            self.imageLabelFBStoryAward.setPixmap(QPixmap(self.rotated_pixmap_award))
            
            
            

                    
    def upload_image(self):
        # Open a QFileDialog to select an image file
        # The getOpenFileName method returns a tuple where the first element is the selected file path
        # The second element is the selected filter, which we don't need, hence the underscore
        bgImage, _ = QFileDialog.getOpenFileName(self, "Open Image", self.image_file, "Image Files (*.png *.jpg *.bmp)")
        if bgImage:
            self.image_file = bgImage

        # If a file was selected (i.e., if self.image_file is not an empty string)
        if self.image_file:
            # Create a QPixmap from the image file
            pixmap = QPixmap(self.image_file)
            # Scale the QPixmap to fit the size of the imageLabel, while keeping the aspect ratio
            pixmap = pixmap.scaled(self.imageLabel.size(), Qt.KeepAspectRatio)
            # Set the QPixmap as the pixmap of the imageLabel
            self.imageLabel.setPixmap(pixmap)
            
    # END OF upload_image METHOD
        
    def change_font(self, font):
        """
        Summary: Changes the font of the labelName QLabel to the selected font.
        
        Purpose: This function is used to change the font of the text in the labelName QLabel. 
        It can be used to customize the appearance of the text in the UI.
        
        Use: It is called with a QFont object as an argument, which represents the new font to be used for the labelName QLabel.
        """
        self.labelName.setFont(font)

    def change_font_size(self, value):
        # Get the current font of the labelName QLabel
        font = self.labelName.font()

        # Set the font size to the value of the dialFontSize
        font.setPointSize(value)

        # Set the font of the labelName QLabel
        self.labelName.setFont(font)
        
        
    def export_multi_pdf(self):
    
        model = self.tableViewExcel.model()

        # Check if model is None
        if model is None:
            print("No model in tableViewExcel.")
            QMessageBox.information(self, "No hay excel", "Es necesario subir un archivo Excel") 
            return

        # Get the number of rows in the model
        num_rows = model.rowCount()

        # Open a dialog for the user to select a directory
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory", "PDF/For print")

        # If the user didn't select a directory, return
        if not folder_path:
            return

        # Iterate over each row
        for row in range(num_rows):
            index = model.index(row, 0)
            cell_text = model.data(index)
            # Set the text of the QLabel
            self.labelName.setText(str(cell_text))
            filename = f"{cell_text}.pdf"
            self.export_pdf(str(cell_text), folder_path, filename, preview=True)
            print(f"Exported {filename} to {folder_path}")
            
    def export_pdf(self, line_content, folder_path=None, filename=None, preview=True):
        # 1. If folder_path is not provided, set it to a default value
        if folder_path is None:
            folder_path = resource_path("PDF\\For print\\")

        # 2. If filename is not provided, open a dialog for the user to select a filename
        if filename is None:
            filename, _ = QFileDialog.getSaveFileName(self, "Save PDF", folder_path, "PDF Files (*.pdf)")

        # 3. If the user didn't select a filename, return
        if not filename:
            return

        # 4. Join the folder path and filename to create the full file path
        pdf_file = os.path.join(folder_path, filename)

        # 5. Ensure directory exists
        os.makedirs(folder_path, exist_ok=True)

        # 6. Validate file path
        if not os.path.isdir(os.path.dirname(pdf_file)):
            print(f"Invalid file path: {pdf_file}")
            return

        # 7. Setup PDF writer
        pdf_writer = QPdfWriter(pdf_file)
        pdf_writer.setPageSizeMM(QSizeF(11 * 25.4, 8.5 * 25.4))
        pdf_writer.setPageMargins(QMarginsF(0, 0, 0, 0))
        pdf_writer.setResolution(300)

        # 8. Setup painter
        painter = QPainter(pdf_writer)
        if not painter.isActive():
            QMessageBox.information(self, "Information", "No puedo sobreescribir un PDF abierto")
            return

        # 9. Draw image
        pixmap = QPixmap(self.image_file)
        if pixmap:
            pixmapWidth, pixmapHeight = pixmap.width(), pixmap.height()
            painter.setViewport(0, 0, pixmapWidth, pixmapHeight)
            painter.setWindow(0, 0, pixmapWidth, pixmapHeight)
            painter.drawPixmap(0, 0, pixmap)
        else:
            print("No pixmap in imageLabel.")
            return

        # 10. Draw text
        font = self.labelName.font()
        font.setPointSize(int(self.labelName.font().pointSize() * 1.5))
        self.labelName.setFont(font)
        font.setPointSize(int(self.labelName.font().pointSize()))
        painter.setFont(font)
        painter.setPen(QColor(Qt.black))

        font_metrics = QFontMetrics(self.labelName.font())
        text_width = int(font_metrics.horizontalAdvance(self.labelName.text()) * 1.55)
        page_center_x = pdf_writer.width() // 2
        page_center_y = int(pdf_writer.height() * .65)

        painter.drawText(page_center_x - text_width, page_center_y, self.labelName.text())

        # 11. Clean up
        painter.end()

        # Only preview the PDF if the preview parameter is True
        if preview:
            self.preview_pdf(pdf_file)

        font.setPointSize(int(self.labelName.font().pointSize() / 1.5))
        self.labelName.setFont(font)
        
    def preview_pdf(self, pdf_file):
        # Open the PDF file in the default PDF viewer
        QDesktopServices.openUrl(QUrl.fromLocalFile(pdf_file))
        # The openUrl method of the QDesktopServices class is used to open the PDF file in the default PDF viewer.
        # The QUrl.fromLocalFile method is used to create a QUrl object from the file path.
        # The openUrl method then opens the PDF file in the default PDF viewer.
        
        
        
if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Load the splash image
    splash_image = QPixmap(resource_path("img\\logo dramzii - Text  only 2.png"))

    # Scale the splash image
    splash_image = splash_image.scaledToWidth(400, Qt.SmoothTransformation)

    # Create a QSplashScreen with the splash image
    splash = QSplashScreen(splash_image)
    splash.show()

    # Create an instance of MainWindow
    window = MainWindow()

    # Use a QTimer to close the splash screen after a delay and start the main window
    QTimer.singleShot(3000, splash.close)
    QTimer.singleShot(3000, window.show)  # Pass the method itself, not the result of calling it

    sys.exit(app.exec_())